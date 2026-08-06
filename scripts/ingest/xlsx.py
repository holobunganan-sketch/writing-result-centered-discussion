from __future__ import annotations

import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from .base import ExtractedUnit, unit


def _fallback(path: Path) -> list[ExtractedUnit]:
    units: list[ExtractedUnit] = []
    with zipfile.ZipFile(path) as archive:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in root.iter():
                if si.tag.rsplit("}", 1)[-1] == "si":
                    shared.append("".join(t.text or "" for t in si.iter() if t.tag.rsplit("}", 1)[-1] == "t"))
        sheets = sorted(n for n in archive.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"))
        for idx, name in enumerate(sheets, 1):
            root = ET.fromstring(archive.read(name))
            for cell in root.iter():
                if cell.tag.rsplit("}", 1)[-1] != "c":
                    continue
                ref = cell.attrib.get("r", "?")
                value = next((x for x in cell.iter() if x.tag.rsplit("}", 1)[-1] == "v"), None)
                text = value.text if value is not None and value.text else ""
                if cell.attrib.get("t") == "s" and text.isdigit() and int(text) < len(shared):
                    text = shared[int(text)]
                if found := unit(f"sheet {idx}!{ref}", text, format="xlsx"):
                    units.append(found)
    return units


def extract_xlsx(path: Path) -> list[ExtractedUnit]:
    try:
        from openpyxl import load_workbook  # type: ignore
        workbook = load_workbook(path, read_only=True, data_only=False)
        units: list[ExtractedUnit] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    if found := unit(f"sheet {sheet.title}!{cell.coordinate}", str(cell.value), format="xlsx", data_type=cell.data_type):
                        units.append(found)
        workbook.close()
        return units or _fallback(path)
    except Exception:
        return _fallback(path)
